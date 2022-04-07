import openpyxl


class ExcelUtils:
    workbook = None
    database = None
    Total_main_database_row = None
    sheet = None
    Total_row = None
    url = None

    def __init__(self):
        ExcelUtils.url = "C:\\Users\\praparihar\\Desktop\\New Microsoft Excel Worksheet.xlsx"
        ExcelUtils.workbook = openpyxl.load_workbook(ExcelUtils.url)
        ExcelUtils.sheet = ExcelUtils.workbook.worksheets[0]
        ExcelUtils.Total_row = ExcelUtils.sheet.max_row
        ExcelUtils.Total_Column = ExcelUtils.sheet.max_column
        ExcelUtils.user_sheet = ExcelUtils.workbook.worksheets[1]
        ExcelUtils.Total_user_row = ExcelUtils.user_sheet.max_row
        ExcelUtils.Total_user_Column = ExcelUtils.user_sheet.max_column
        ExcelUtils.database = ExcelUtils.workbook.worksheets[2]
        ExcelUtils.Total_main_database_row = ExcelUtils.database.max_row
        ExcelUtils.Total_main_database_Column = ExcelUtils.database.max_column





